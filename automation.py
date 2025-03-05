import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import re
from datetime import datetime

def calculate_hours(session_time):
    try:
        # Extract start and end times using regex
        match = re.search(r'(\d{1,2}:\d{2}\s*[APap][Mm])\s*-\s*(\d{1,2}:\d{2}\s*[APap][Mm])', str(session_time))
        if match:
            start_time = datetime.strptime(match.group(1), "%I:%M %p")
            end_time = datetime.strptime(match.group(2), "%I:%M %p")
            if end_time < start_time:
                end_time += pd.Timedelta(days=1)  # Handle overnight sessions
            duration = (end_time - start_time).total_seconds() / 3600  # Convert seconds to hours
            return round(duration, 2)
    except:
        return None
    return None

def transfer_and_update_data(source_file, target_file, sheet_name):
    columns = ["Date", "Day", "Session Time", "Client", "Program Name", "Batch", "Session Name",
               "Mentor / Faculty", "Year", "Month", "Vendor Name", "No.of Hours", "Remarks"]

    # Read source Excel file
    df = pd.read_excel(source_file, usecols=columns, dtype=str)

    # Convert Date column to datetime format and extract proper values
    df["Date"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime('%Y-%m-%d')
    df["Day"] = pd.to_datetime(df["Date"], errors='coerce').dt.day_name()
    df["Year"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime('%Y')
    df["Month"] = pd.to_datetime(df["Date"], errors='coerce').dt.strftime('%B')
    df["Session Time"] = df["Session Time"].apply(lambda x: re.sub(r'\s*to\s*', ' - ', str(x)))

    # Sort data by Date and Session Time
    df = df.sort_values(by=["Date", "Session Time"], ascending=[True, True])

    # Add Serial Number Column
    df.insert(0, "S. No", range(1, len(df) + 1))

    # Update No.of Hours if it's missing or zero
    df["No.of Hours"] = df.apply(lambda row: calculate_hours(row["Session Time"]) if pd.isna(row["No.of Hours"]) or row["No.of Hours"] in [0, "0", ""] else row["No.of Hours"], axis=1)

    # Check if target file exists
    try:
        wb = load_workbook(target_file)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(target_file)  # Create file if not exists

    # Check if the sheet exists, if not create one
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Save changes and close workbook
    wb.save(target_file)

    # Load updated workbook again with Pandas
    with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Open workbook again to apply formatting
    wb = load_workbook(target_file)
    ws = wb[sheet_name]

    # Auto-fit column width and row height
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    # Apply orange fill to header row
    fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill

    # Save final changes
    wb.save(target_file)
    print(f"Data transferred and formatted successfully to {target_file}.")

# Example usage - Updated with actual file paths
source_file = "C:/Users/asus/OneDrive - NIIT Limited/Automation/Source_data.xlsx"
target_file = "C:/Users/asus/OneDrive - NIIT Limited/Automation/Destination_data.xlsx"
sheet_name = "Formatted Data"

transfer_and_update_data(source_file, target_file, sheet_name)
